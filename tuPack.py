import csv
import openpyxl
import tkinter as tk
from tkinter import filedialog, ttk
import warnings
import logging
import os
from datetime import datetime

# Log-Datei einrichten
def setup_logging(csv_path):
    # Hole das aktuelle Datum im gewünschten Format
    datum = datetime.now().strftime("%Y-%m-%d")

    # Erstelle den Dateipfad mit dem aktuellen Datum
    log_path = os.path.join(os.path.dirname(csv_path), f"Abweichungsmeldungen_daten_log_{datum}.log")
    logging.basicConfig(
        filename=log_path,  # Log-Datei im gleichen Verzeichnis wie die CSV-Datei
        level=logging.INFO,  # Alle Log-Nachrichten mit dem Level INFO oder höher werden gespeichert
        format="%(asctime)s - %(levelname)s - %(message)s",  # Log-Nachrichtenformat
    )

# Warnungen ignorieren
warnings.simplefilter("ignore", category=UserWarning)
# Funktion zum Auswählen der Excel-Dateien (Mehrfachauswahl)
def select_excel_files():
    global selected_files  # Zugriff auf die globale Variable
    file_paths = filedialog.askopenfilenames(
        title="Wählen Sie die Excel-Dateien aus",
        filetypes=[("Excel Dateien", "*.xlsx;*.xlsm")]

    )

    if file_paths:
        # Hole den Ordnerpfad der ersten ausgewählten Datei (alle Dateien befinden sich im gleichen Ordner)

        folder_path = os.path.dirname(file_paths[0])

        datei_anzahl = len(file_paths)  # Hier verwenden wir file_paths, nicht selected_files
        excel_files_label.config(text=f"Es wurden {datei_anzahl} Excel-Dateien für die Verarbeitung ausgewählt.")
        print(f"Ordner mit Excel-Dateien: {folder_path}")  # Detailierte Ausgabe

        # Im Hintergrund alle vollständigen Dateipfade speichern
        selected_files = file_paths  # Speichere die ausgewählten Dateien in der globalen Variable


# Funktion zum Auswählen des Speicherorts und Namens der CSV-Datei
def select_csv_path():
    folder_path = filedialog.askdirectory(title="Ordner zum Speichern der CSV auswählen")
    if folder_path:
        # Benenne die CSV-Datei nach dem aktuellen Datum
        date_str = datetime.now().strftime("%Y-%m-%d")
        csv_filename = f"Abweichungsmeldungen_daten_{date_str}.csv"
        full_path = os.path.join(folder_path, csv_filename)
        csv_file_label.config(text=f"CSV wird gespeichert unter: {full_path}")
        print(f"CSV-Datei wird gespeichert unter: {full_path}")  # Detailierte Ausgabe
        return full_path
    else:
        csv_file_label.config(text="Kein Ordner ausgewählt.")
        return None

# Globale Variable zum Speichern der ausgewählten Excel-Dateipfade
selected_files = []



def process_excel_files():
    global selected_files  # Zugriff auf die globale Variable

    # Überprüfen, ob Excel-Dateien ausgewählt wurden
    if not selected_files:
        result_label.config(text="Bitte wählen Sie Excel-Dateien aus.")
        print("Fehler: Keine Excel-Dateien ausgewählt.")  # Detailierte Ausgabe
        return

    if not csv_file_label.cget("text") or csv_file_label.cget("text") == "Kein Ordner ausgewählt.":
        result_label.config(text="Bitte wählen Sie einen Speicherort für die CSV-Datei aus.")
        print("Fehler: Kein Speicherort für die CSV-Datei ausgewählt.")  # Detailierte Ausgabe
        return

    # Holen des Speicherorts der CSV-Datei
    csv_path = csv_file_label.cget("text").replace("CSV wird gespeichert unter: ", "")
    setup_logging(csv_path)  # Log-Datei im gleichen Ordner wie CSV setzen
    print(f"Verarbeite {len(selected_files)} Excel-Dateien...")  # Detailierte Ausgabe

    # Gesamtanzahl der Dateien
    total_files = len(selected_files)

    # Fortschrittsbalken zurücksetzen
    progress_bar["value"] = 0
    progress_bar["maximum"] = total_files

    # Prozentanzeige initialisieren
    percentage_label.config(text="0%")

    with open(csv_path, mode='w', newline='', encoding='utf-8-sig') as file:
        writer = csv.writer(file)
        headers = ["Datum", "MeldungsNr", "UHPT-Bestellnummer und Pos.", "UHPT-Auftragsnummer",
                   "Kaufm. Sachbearbeiter UHPT", "Techn. Sachbearbeiter UHPT",
                   "UHPT-Zeichnungsnummer und Revisionstand",
                   "Benennung", "Name und Anschrift Lieferant", "Kontaktdaten Lieferant", "Auftragsnummer Lieferant",
                   "Bestellmenge", "Betroffene Menge (n.i.O.)", "Fehlerbeschreibung", "Anlagenverzeichnis",
                   "Vorgeschlagene Maßnahme / Sofortmaßnahme", "Fehlerursache", "Korrekturmaßnahme zur Fehlerursache",
                   "Sonderfreigabe / Tolerierung","Nacharbeit (Lieferant)", "Nacharbeit (UHDE)", "Ausschuss (Verschrottung Lieferant)",
                   "Ausschuss (Verschrottung Uhde)",  "Auftrag reduzieren", "Neufertigung UHPT",
                   "Nacharbeitskosten zu Lasten UHPT", "Neufertigung Lieferant (Materialbestellung nicht erforderlich)",
                   "Neufertigung Lieferant (Materialbestellung erforderlich)", "Nacharbeitskosten zu Lasten Lieferant"]
        writer.writerow(headers)

        for idx, file_path in enumerate(selected_files, start=1):
            try:
                print(f"Verarbeite Datei {idx}/{total_files}: {file_path}")  # Detailierte Ausgabe
                logging.info(f"Verarbeite Datei {idx}/{total_files}: {file_path}")

                # Arbeitsbuch und Blätter laden
                wb = openpyxl.load_workbook(file_path)  # Verwende die Variable file_path, die die ausgewählte Datei enthält
                sheet1 = wb['Arbeitsblatt1']  # Ersetze 'Sheet1' mit dem tatsächlichen Namen des ersten Blattes
                sheet2 = wb['Arbeitsblatt2']  # Ersetze 'Sheet2' mit dem tatsächlichen Namen des zweiten Blattes

                # Werte aus dem ersten Arbeitsblatt extrahieren
                datum = sheet1['A26'].value
                datum = datum.split(": ")[1]
                bestellnummer_pos = sheet1['C5'].value  # Wert aus C5
                auftragsnummer = sheet1['C6'].value  # Wert aus C6
                sachbearbeiter = sheet1['C7'].value  # Wert aus C7
                techn_sachbearbeiter = sheet1['C8'].value  # Wert aus C8
                zeichnungsnummer = sheet1['C9'].value  # Wert aus C9
                benennung = sheet1['C10'].value  # Wert aus C10

                # Wert aus den verbundenen Zellen I5 bis M6 (Name und Anschrift Lieferant)
                lieferant_name_anschrift = sheet1['I5'].value  # Wert aus der ersten verbundenen Zelle
                lieferant_name_anschrift = lieferant_name_anschrift.replace('\n', ' ')  # Zeilenumbrüche entfernen

                # Wert aus den verbundenen Zellen I7 bis N8 (Kontaktdaten Lieferant)
                lieferant_kontaktdaten = sheet1['I7'].value  # Wert aus der ersten verbundenen Zelle
                lieferant_kontaktdaten = lieferant_kontaktdaten.replace('\n', ' ')  # Zeilenumbrüche entfernen

                # Weitere Werte extrahieren
                auftragsnummer_lieferant = sheet1['I9'].value  # Wert aus I9 bis N9
                bestellmenge = sheet1['I10'].value  # Wert aus IJK10
                betroffene_menge = sheet1['L10'].value  # Wert aus LMN10

                # Fehlerbeschreibung aus den Zellen C12 bis N12 extrahieren
                fehlerbeschreibung = ''.join(
                    [str(sheet1.cell(row=12, column=col).value or '') for col in range(3, 15)])  # C12 bis N12
                fehlerbeschreibung = fehlerbeschreibung.replace('\n', ' ')  # Zeilenumbrüche entfernen

                # Anlagenverzeichnis aus den Zellen C13 bis N13 extrahieren
                anlagenverzeichnis = ''.join(
                    [str(sheet1.cell(row=13, column=col).value or '') for col in range(3, 15)])  # C13 bis N13
                anlagenverzeichnis = anlagenverzeichnis.replace('\n', ' ')  # Zeilenumbrüche entfernen
#
                # Vorgeschlagene Maßnahme / Sofortmaßnahme aus C14 bis N14 extrahieren
                vorgeschlagene_massnahme = ''.join(
                    [str(sheet1.cell(row=14, column=col).value or '') for col in range(3, 15)])  # C14 bis N14
                vorgeschlagene_massnahme = vorgeschlagene_massnahme.replace('\n', ' ')  # Zeilenumbrüche entfernen

                # Fehlerursache aus den Zellen C15 bis N15 extrahieren
                fehlerursache = ''.join(
                    [str(sheet1.cell(row=15, column=col).value or '') for col in range(3, 15)])  # C15 bis N15
                fehlerursache = fehlerursache.replace('\n', ' ')  # Zeilenumbrüche entfernen

                # Korrekturmaßnahme zur Fehlerursache aus den Zellen C16 bis N16 extrahieren
                korrekturmassnahme = ''.join(
                    [str(sheet1.cell(row=16, column=col).value or '') for col in range(3, 15)])  # C16 bis N16
                korrekturmassnahme = korrekturmassnahme.replace('\n', ' ')  # Zeilenumbrüche entfernen

                # Sonderfreigabe / Tolerierung aus Zelle E21 extrahieren
                sonderfreigabe = sheet1['E21'].value  # Wert aus E21

                # Nacharbeit (Lieferant) aus Zelle E22 extrahieren
                nacharbeit_lieferant = sheet1['E22'].value  # Wert aus E22

                # Nacharbeit (UHDE) aus Zelle E23 extrahieren
                nacharbeit_uhde = sheet1['E23'].value  # Wert aus E23

                # Ausschuss (Verschrottung Lieferant) aus den verbundenen Zellen M21 und N21 extrahieren
                ausschuss_verschrottung_lieferant = sheet1['M21'].value  # Wert aus M21 (und verbunden mit N21)

                # Ausschuss (Verschrottung Uhde) aus den verbundenen Zellen M21 und N21 extrahieren (Wiederholung)
                ausschuss_verschrottung_uhde = sheet1['M22'].value  # Wert aus M22 (und verbunden mit N21)

                # Neue Werte aus dem zweiten Arbeitsblatt extrahieren
                auftrag_reduzieren = sheet2['D5'].value  # Wert aus D5
                neufertigung_uhpt = sheet2['D6'].value  # Wert aus D6
                nacharbeitskosten_zu_lasten_uhpt = sheet2['D7'].value  # Wert aus D7
                neufertigung_lieferant_materialbestellung_nicht_erforderlich = sheet2['I5'].value  # Wert aus I5
                neufertigung_lieferant_materialbestellung_erforderlich = sheet2['I6'].value  # Wert aus I6
                nacharbeitskosten_zu_lasten_lieferant = sheet2['I7'].value  # Wert aus I7

                # Meldungsnummer aus den verbundenen Zellen K2 bis N2 extrahieren
                meldungsnummer = sheet1['K2'].value  # Wert aus K2
                meldungsnummer = meldungsnummer.replace(' ', '')
                meldungsnummer = meldungsnummer.replace('\n', '')  # Zeilenumbrüche entfernen

                row = [datum, meldungsnummer, bestellnummer_pos, auftragsnummer, sachbearbeiter, techn_sachbearbeiter,
                       zeichnungsnummer, benennung, lieferant_name_anschrift, lieferant_kontaktdaten,
                       auftragsnummer_lieferant, bestellmenge, betroffene_menge, fehlerbeschreibung,
                       anlagenverzeichnis, vorgeschlagene_massnahme, fehlerursache, korrekturmassnahme,
                       sonderfreigabe, nacharbeit_lieferant, nacharbeit_uhde, ausschuss_verschrottung_lieferant, ausschuss_verschrottung_uhde,
                       auftrag_reduzieren, neufertigung_uhpt, nacharbeitskosten_zu_lasten_uhpt,
                       neufertigung_lieferant_materialbestellung_nicht_erforderlich,
                       neufertigung_lieferant_materialbestellung_erforderlich, nacharbeitskosten_zu_lasten_lieferant ]
                writer.writerow(row)

                # Fortschritt anzeigen
                progress_bar["value"] = idx
                percentage_label.config(text=f"{int((idx / total_files) * 100)}%")
                root.update_idletasks()  # GUI aktualisieren

            except Exception as e:
                logging.error(f"Fehler bei der Verarbeitung von Datei {file_path}: {e}")
                print(f"Fehler bei Datei {file_path}: {e}")

    result_label.config(text="Daten wurden erfolgreich exportiert.")
    print(f"Daten wurden erfolgreich exportiert nach {csv_path}")

# GUI Setup
root = tk.Tk()
root.title("TuPack Version 0.1 - Abweichungsberichte extrahieren")


# Modernes Designc
root.geometry("600x400")
root.configure(bg="#f4f4f9")

# UI Elemente
select_button = tk.Button(root, text="Excel-Dateien auswählen", command=select_excel_files, bg="#4CAF50", fg="white",
                          font=("Arial", 12), relief="flat")
select_button.pack(pady=10, fill="x")

excel_files_label = tk.Label(root, text="Keine Dateien ausgewählt", wraplength=500, bg="#f4f4f9", font=("Arial", 10))
excel_files_label.pack(pady=5)

select_csv_button = tk.Button(root, text="Speicherort der CSV-Datei auswählen", command=select_csv_path, bg="#4CAF50", fg="white",
                              font=("Arial", 12), relief="flat")
select_csv_button.pack(pady=10, fill="x")

csv_file_label = tk.Label(root, text="Kein Ordner ausgewählt.", bg="#f4f4f9", font=("Arial", 10))
csv_file_label.pack(pady=5)

process_button = tk.Button(root, text="Verarbeitung starten", command=process_excel_files, bg="#008CBA", fg="white",
                           font=("Arial", 12), relief="flat")
process_button.pack(pady=20, fill="x")

# Fortschrittsbalken und Prozentanzeige
progress_bar = ttk.Progressbar(root, length=400, mode="determinate", style="TProgressbar")
progress_bar.pack(pady=10)

percentage_label = tk.Label(root, text="0%", bg="#f4f4f9", font=("Arial", 12))
percentage_label.pack(pady=5)

result_label = tk.Label(root, text="", bg="#f4f4f9", font=("Arial", 12))
result_label.pack(pady=10)

# Credits Label

credits_label = tk.Label(root, text="Entwickelt von Karim Abdulhadi | Version 0.1 Beta", bg="#f4f4f9", font=("Arial", 12, "italic"))
credits_label.pack(side="bottom", pady=10)


# Stil für Fortschrittsbalken
style = ttk.Style()
style.configure("TProgressbar",
                thickness=30,
                length=400,
                mode="determinate",
                maximum=100,
                )


root.mainloop()
